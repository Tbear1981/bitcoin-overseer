from openpyxl import Workbook

# create a dummy list of texts to write to excel file
divs = [[chr(i)*8, chr(i+1)*8] for i in range(65, 75, 1)]

wb = Workbook()             # open new workbook, use load_workbook if existing
ws = wb.create_sheet(title="Example")
for div in divs:
    row = [div[0], div[1]]  # construct a row: shown only for example purposes
    ws.append(row)          # could use ws.append(div) since each div is a list 
wb.save('example.xlsx')     # save workbook, will overwrite if exists

